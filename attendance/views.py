from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from .forms import PersonForm
from .models import PersonProfile, QRData, Attendance
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponseBadRequest
import datetime
from django.utils.timezone import now
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.dateparse import parse_datetime
from datetime import datetime, timedelta
from collections import defaultdict, deque
from django.contrib import messages
import openpyxl
from django.http import HttpResponse
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .serializers import AttendanceSerializer
from rest_framework import generics
from .serializers import PersonProfileSerializer
from rest_framework import status
from django.utils import timezone
from django.db.models import Sum, Count, F, ExpressionWrapper, DurationField
from dateutil import parser
from datetime import timedelta, datetime


# User Register  
@login_required
def user_register(request):
    form = UserCreationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        login(request, user)
        return redirect("dashboard")
    return render(request, "attendance/register.html", {"form": form})

# User Login 
def user_login(request):
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == "POST" and form.is_valid():
        login(request, form.get_user())
        return redirect("qr_scanner")
    return render(request, "attendance/login.html", {"form": form})


# User Logout
def user_logout(request):
    logout(request)
    return redirect("login")

# Dashboard
@login_required
def dashboard(request):
    # persons = PersonProfile.objects.filter(user=request.user) // User တစ်ယောက်ချင်းစီ Create လုပ်ထားရေကောင်ကို ပြစွာ ကိုယ် create လုပ်ထားရေကောင်ကို ယာ မြင်ရဖို့
    persons = PersonProfile.objects.all()
    return render(request, "attendance/dashboard.html", {"persons": persons})

# Add Person
@login_required
def add_person(request):
    if request.method == "POST":
        form = PersonForm(request.POST)
        if form.is_valid():
            person = form.save(commit=False)
            person.user = request.user
            person.save()
            QRData.objects.create(person=person)
            return redirect("dashboard")
    else:
        form = PersonForm()
    return render(request, "attendance/add_person.html", {"form": form})

# QR Scanner or Home Page
# @login_required
def qr_scanner(request):
    # QR Scanner page ကို render ပြန်ပေးမယ်
    return render(request, "attendance/qr_scanner.html")

# Save to Attandance when Qr is scanned 
@csrf_exempt
def scan_qr(request):
    if request.method == 'POST':
        qr_data = request.POST.get('qr_data')
        try:
            person_id, name = qr_data.split(',')
            person = PersonProfile.objects.get(id=person_id, name=name)

            # Last attendance record for this person (latest timestamp)
            last_record = Attendance.objects.filter(person=person).order_by('-timestamp').first()

            # Determine next action based on last record
            if not last_record or last_record.action == 'out':
                action = 'in'  # next is check-in
            else:
                action = 'out'  # next is check-out

            # Create new attendance record
            Attendance.objects.create(person=person, timestamp=now(), action=action)

            return JsonResponse({'status': 'success', 'action': f'Check-{action}'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return HttpResponseBadRequest("Only POST method allowed.")


# Attendance List
def attendance_list(request):
    attendances = Attendance.objects.select_related('person').all()


    
    q = request.GET.get("q", "")
    start = request.GET.get("start_timestamp", "")
    end = request.GET.get("end_timestamp", "")
    action = request.GET.get("action", "")

    if q:
        attendances = attendances.filter(person__name__icontains=q)

    
    if start:
        attendances = attendances.filter(timestamp__gte=parse_datetime(start))
    if end:
        attendances = attendances.filter(timestamp__lte=parse_datetime(end))

    if action == "in":
        attendances = attendances.filter(action="in")
    elif action == "out":
        attendances = attendances.filter(action="out")

    paginator = Paginator(attendances.order_by('-timestamp'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "attendance/attendance_list.html", {
        "attendances": page_obj,
    })

# တစ်နိကို တစ်ယောက်ယာ သိမ်းချင်ကေ သုံး
# # views.py
# from django.shortcuts import render
# from collections import defaultdict
# from datetime import timedelta, datetime
# from .models import Attendance
# from django.core.paginator import Paginator
# from django.db.models import Q

# from django.utils import timezone

# def paired_attendance_list(request):
#     all_records = Attendance.objects.select_related('person').order_by('person_id', 'date', 'timestamp')
#     mm_time = timezone.localtime(timezone.now())
#     print(mm_time)
#     print("hello")


#     # Filters
#     action_filter = request.GET.get('action')
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     search_query = request.GET.get('q')

#     if action_filter:
#         all_records = all_records.filter(action=action_filter)

#     if start_date:
#         all_records = all_records.filter(date__gte=start_date)

#     if end_date:
#         all_records = all_records.filter(date__lte=end_date)

#     if search_query:
#         all_records = all_records.filter(Q(person__name__icontains=search_query))

#     # Pairing logic
#     pairs = defaultdict(dict)

#     for record in all_records:
#         key = (record.person.id, record.date)
#         if record.action == 'in':
#             pairs[key]['in'] = record.timestamp
#         elif record.action == 'out':
#             pairs[key]['out'] = record.timestamp
#         pairs[key]['person'] = record.person
#         pairs[key]['date'] = record.date

#     results = []
#     for (person_id, date), data in pairs.items():
#         checkin = data.get('in')
#         checkout = data.get('out')
#         duration = None
#         if checkin and checkout:
#             duration = checkout - checkin
#         results.append({
#             'person': data['person'],
#             'date': date,
#             'checkin': checkin,
#             'checkout': checkout,
#             'duration': duration
#         })

#     paginator = Paginator(results, 10)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)

#     return render(request, 'attendance/paired_attendance.html', {
#         'records': page_obj,
#         'action_filter': action_filter,
#         'page_obj': page_obj,
#         'start_date': start_date,
#         'end_date': end_date,
#         'search_query': search_query
#     })

# Format Duration for customization time
def format_duration(td):
    if not td:
        return None
    total_seconds = int(td.total_seconds())  # ဒသမမပါစေ
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours}:{minutes:02}:{seconds:02}"

# def paired_attendance_list(request):
#     # Get all records ordered by person/date/timestamp
#     all_records = Attendance.objects.select_related('person').order_by('person_id', 'date', 'timestamp')

#     # Filters
#     action_filter = request.GET.get('action')
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     search_query = request.GET.get('q')

#     if action_filter:
#         all_records = all_records.filter(action=action_filter)

#     if start_date:
#         all_records = all_records.filter(date__gte=start_date)

#     if end_date:
#         all_records = all_records.filter(date__lte=end_date)

#     if search_query:
#         all_records = all_records.filter(person__name__icontains=search_query)

#     # Group by person + date
#     grouped = defaultdict(list)
#     for record in all_records:
#         key = (record.person_id, record.date)
#         grouped[key].append(record)

#     results = []

#     # Pairing algorithm
#     for (person_id, date), records in grouped.items():
#         records = sorted(records, key=lambda r: r.timestamp)
#         pending_checkins = deque()

#         for rec in records:
#             if rec.action == 'in':
#                 # Append checkin to queue
#                 pending_checkins.append(rec.timestamp)
#             elif rec.action == 'out':
#                 if pending_checkins:
#                     # Pop the oldest checkin to pair
#                     checkin_time = pending_checkins.popleft()
#                     duration = rec.timestamp - checkin_time
#                     results.append({
#                         'person': rec.person,
#                         'date': date,
#                         'checkin': checkin_time,
#                         'checkout': rec.timestamp,
#                         'duration': format_duration(duration)
#                     })
#                 else:
#                     # Checkout without checkin
#                     results.append({
#                         'person': rec.person,
#                         'date': date,
#                         'checkin': None,
#                         'checkout': rec.timestamp,
#                         'duration': None
#                     })

#         # Remaining checkins without checkout
#         while pending_checkins:
#             checkin_time = pending_checkins.popleft()
#             results.append({
#                 'person': rec.person,
#                 'date': date,
#                 'checkin': checkin_time,
#                 'checkout': None,
#                 'duration': None
#             })

#     # Pagination
#     paginator = Paginator(results, 10)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)

#     return render(request, 'attendance/paired_attendance.html', {
#         'records': page_obj,
#         'page_obj': page_obj,
#         'action_filter': action_filter,
#         'start_date': start_date,
#         'end_date': end_date,
#         'search_query': search_query
#     })


# Paired Attendance List (Check in / out Paired) or show data with duration
def paired_attendance_list(request):
    date_str = request.GET.get("date")
    name_query = request.GET.get("name", "").strip()
    action_filter = request.GET.get("action", "").strip().lower()

    today = timezone.now().date()

    if date_str:
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            selected_date = today
    else:
        selected_date = today

    records = Attendance.objects.filter(
        date=selected_date
    ).select_related("person").order_by("person_id", "timestamp")

    if name_query:
        records = records.filter(person__name__icontains=name_query)

    if action_filter in ["in", "out"]:
        records = records.filter(action=action_filter)

#    Group by person + date
    grouped = defaultdict(list)
    for record in records:
        key = (record.person_id, record.date)
        grouped[key].append(record)

    results = []

    # Pairing algorithm
    for (person_id, date), records in grouped.items():
        records = sorted(records, key=lambda r: r.timestamp)
        pending_checkins = deque()

        for rec in records:
            if rec.action == 'in':
                # Append checkin to queue
                pending_checkins.append(rec.timestamp)
            elif rec.action == 'out':
                if pending_checkins:
                    # Pop the oldest checkin to pair
                    checkin_time = pending_checkins.popleft()
                    duration = rec.timestamp - checkin_time
                    results.append({
                        'person': rec.person,
                        'date': date,
                        'checkin': checkin_time,
                        'checkout': rec.timestamp,
                        'duration': format_duration(duration)
                    })
                else:
                    # Checkout without checkin
                    results.append({
                        'person': rec.person,
                        'date': date,
                        'checkin': None,
                        'checkout': rec.timestamp,
                        'duration': None
                    })

        # Remaining checkins without checkout
        while pending_checkins:
            checkin_time = pending_checkins.popleft()
            results.append({
                'person': rec.person,
                'date': date,
                'checkin': checkin_time,
                'checkout': None,
                'duration': None
            })

    # Pagination
    paginator = Paginator(results, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'attendance/paired_attendance.html', {
        'records': page_obj,
        'page_obj': page_obj,
        'action_filter': action_filter,
        # 'selected_date': today,
        'selected_date': selected_date,
        # 'end_date': end_date,
        'name_query': name_query
    })

# Edit Person
def edit_person(request, pk):
    person = get_object_or_404(PersonProfile, pk=pk)
    if request.method == 'POST':
        form = PersonForm(request.POST, instance=person)
        if form.is_valid():
            form.save()
            messages.success(request, "Person updated successfully.")
            return redirect('dashboard')
    else:
        form = PersonForm(instance=person)
    return render(request, 'attendance/edit_person.html', {'form': form})

# Delete  Person
def delete_person(request, pk):
    person = get_object_or_404(PersonProfile, pk=pk)
    if request.method == 'POST':
        person.delete()
        messages.success(request, "Person deleted successfully.")
        return redirect('dashboard')
    return render(request, 'attendance/delete_person_confirm.html', {'person': person})



# Export Attendance Excel
def export_attendance_excel(request):
    # date_str = request.GET.get("date")
    # today = timezone.now().date()

    # if date_str:
    #     try:
    #         selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    #     except ValueError:
    #         selected_date = today
    # else:
    #     selected_date = today

    # all_records = Attendance.objects.filter(
    #     date=selected_date
    # ).select_related("person").order_by("timestamp")
    # wb = openpyxl.Workbook()
    # ws = wb.active
    # ws.title = "Attendance Records"

    # # Table Header
    # ws.append(['No','Name', 'Date', 'Check-In', 'Check-Out', 'Duration'])

    

    # # all_records = Attendance.objects.select_related('person').order_by('person_id', 'date', 'timestamp')
    # grouped = defaultdict(list)
    # for record in all_records:
    #     key = (record.person.id, record.date)
    #     grouped[key].append(record)

    # results = []

    # for (person_id, date), records in grouped.items():
    #     records = sorted(records, key=lambda r: r.timestamp)
    #     pending_checkins = deque()

    #     for rec in records:
    #         if rec.action == 'in':
    #             pending_checkins.append(rec.timestamp)
    #         elif rec.action == 'out':
    #             if pending_checkins:
    #                 checkin_time = pending_checkins.popleft()
    #                 duration = rec.timestamp - checkin_time
    #                 results.append({
    #                     'person': rec.person,
    #                     'date': date,
    #                     'checkin': checkin_time,
    #                     'checkout': rec.timestamp,
    #                     'duration': duration
    #                 })
    #             else:
    #                 results.append({
    #                     'person': rec.person,
    #                     'date': date,
    #                     'checkin': None,
    #                     'checkout': rec.timestamp,
    #                     'duration': None
    #                 })

    #     while pending_checkins:
    #         checkin_time = pending_checkins.popleft()
    #         results.append({
    #             'person': rec.person,
    #             'date': date,
    #             'checkin': checkin_time,
    #             'checkout': None,
    #             'duration': None
    #         })

    # for idx, row in enumerate(results, start=1):
    #     ws.append([
    #         idx,
    #         row['person'].name,
    #         row['date'].strftime('%Y-%m-%d'),
    #         row['checkin'].strftime('%H:%M:%S') if row['checkin'] else '',
    #         row['checkout'].strftime('%H:%M:%S') if row['checkout'] else '',
    #         str(row['duration']) if row['duration'] else '',
    #     ])

    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename=attendance.xlsx'
    # wb.save(response)
    # return response
    date_str = request.GET.get("date")
    today = timezone.now().date()

    if date_str:
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Invalid date", status=400)
    else:
        return HttpResponse("Date is required", status=400)

    records = Attendance.objects.filter(
        date=selected_date
    ).select_related("person").order_by("person_id", "timestamp")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paired Attendance"

    ws.append(["No", "Name", "Check-in", "Checkout", "Duration"])

    grouped = defaultdict(list)
    for r in records:
        grouped[r.person].append(r)

    row_idx = 1
    for person, recs in grouped.items():
        recs = sorted(recs, key=lambda r: r.timestamp)
        pending_checkins = deque()
        for r in recs:
            if r.action == "in":
                pending_checkins.append(r.timestamp)
            elif r.action == "out":
                if pending_checkins:
                    checkin_time = pending_checkins.popleft()
                    duration = r.timestamp - checkin_time
                    ws.append([
                        row_idx,
                        person.name,
                        checkin_time.strftime("%Y-%m-%d %H:%M:%S"),
                        r.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                        format_duration(duration)
                    ])
                    row_idx +=1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=paired_attendance_{selected_date}.xlsx"
    wb.save(response)
    return response


# Export Attendance PDF
def export_attendance_pdf(request):

    # response = HttpResponse(content_type='application/pdf')
    # response['Content-Disposition'] = 'attachment; filename=attendance.pdf'

    # p = canvas.Canvas(response, pagesize=A4)
    # width, height = A4
    # y = height - 50
    # pdfmetrics.registerFont(TTFont('Myanmar', 'fonts/Pyidaungsu.ttf'))
    # p.setFont("Myanmar", 16)
    # p.drawString(50, y, "Attendance Records")
    # y -= 30

    # p.setFont("Myanmar", 10)
    # p.drawString(50, y, "No")
    # p.drawString(80, y, "Name")
    # p.drawString(180, y, "Date")
    # p.drawString(260, y, "Check-In")
    # p.drawString(340, y, "Check-Out")
    # p.drawString(430, y, "Duration")
    # y -= 20

    # date_str = request.GET.get("date")
    # today = timezone.now().date()

    # if date_str:
    #     try:
    #         selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    #     except ValueError:
    #         selected_date = today
    # else:
    #     selected_date = today

    # all_records = Attendance.objects.filter(
    #     date=selected_date
    # ).select_related("person").order_by("timestamp")
    # wb = openpyxl.Workbook()
    # ws = wb.active
    # ws.title = "Attendance Records"

    # # Table Header
    # ws.append(['No','Name', 'Date', 'Check-In', 'Check-Out', 'Duration'])

    # grouped = defaultdict(list)
    # for record in all_records:
    #     key = (record.person.id, record.date)
    #     grouped[key].append(record)

    # results = []

    # for (person_id, date), records in grouped.items():
    #     records = sorted(records, key=lambda r: r.timestamp)
    #     pending_checkins = deque()

    #     for rec in records:
    #         if rec.action == 'in':
    #             pending_checkins.append(rec.timestamp)
    #         elif rec.action == 'out':
    #             if pending_checkins:
    #                 checkin_time = pending_checkins.popleft()
    #                 duration = rec.timestamp - checkin_time
    #                 results.append({
    #                     'person': rec.person,
    #                     'date': date,
    #                     'checkin': checkin_time,
    #                     'checkout': rec.timestamp,
    #                     'duration': duration
    #                 })
    #             else:
    #                 results.append({
    #                     'person': rec.person,
    #                     'date': date,
    #                     'checkin': None,
    #                     'checkout': rec.timestamp,
    #                     'duration': None
    #                 })

    #     while pending_checkins:
    #         checkin_time = pending_checkins.popleft()
    #         results.append({
    #             'person': rec.person,
    #             'date': date,
    #             'checkin': checkin_time,
    #             'checkout': None,
    #             'duration': None
    #         })

    # for idx, row in enumerate(results, start=1):
    #     if y < 50:
    #         p.showPage()
    #         y = height - 50
    #     p.drawString(50, y, str(idx))
    #     p.drawString(80, y, row['person'].name)
    #     p.drawString(180, y, row['date'].strftime('%Y-%m-%d'))
    #     p.drawString(260, y, row['checkin'].strftime('%H:%M:%S') if row['checkin'] else '-')
    #     p.drawString(340, y, row['checkout'].strftime('%H:%M:%S') if row['checkout'] else '-')
    #     p.drawString(430, y, str(row['duration']) if row['duration'] else '-')
    #     y -= 18

    # p.save()
    # return response

    date_str = request.GET.get("date")
    today = timezone.now().date()

    if date_str:
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Invalid date", status=400)
    else:
        return HttpResponse("Date is required", status=400)

    records = Attendance.objects.filter(
        date=selected_date
    ).select_related("person").order_by("person_id", "timestamp")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f"attachment; filename=paired_attendance_{selected_date}.pdf"

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 50

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, f"Paired Attendance - {selected_date}")
    y -= 30

    p.setFont("Helvetica", 10)
    p.drawString(50, y, "No")
    p.drawString(80, y, "Name")
    p.drawString(180, y, "Date")
    p.drawString(260, y, "Check-In")
    p.drawString(340, y, "Check-Out")
    p.drawString(430, y, "Duration")
    y -= 15

    grouped = defaultdict(list)
    for record in records:
        key = (record.person.id, record.date)
        grouped[key].append(record)

    results = []

    for (person_id, date), records in grouped.items():
        records = sorted(records, key=lambda r: r.timestamp)
        pending_checkins = deque()

        for rec in records:
            if rec.action == 'in':
                pending_checkins.append(rec.timestamp)
            elif rec.action == 'out':
                if pending_checkins:
                    checkin_time = pending_checkins.popleft()
                    duration = rec.timestamp - checkin_time
                    results.append({
                        'person': rec.person,
                        'date': date,
                        'checkin': checkin_time,
                        'checkout': rec.timestamp,
                        'duration': duration
                    })
                else:
                    results.append({
                        'person': rec.person,
                        'date': date,
                        'checkin': None,
                        'checkout': rec.timestamp,
                        'duration': None
                    })

        while pending_checkins:
            checkin_time = pending_checkins.popleft()
            results.append({
                'person': rec.person,
                'date': date,
                'checkin': checkin_time,
                'checkout': None,
                'duration': None
            })

    for idx, row in enumerate(results, start=1):
        if y < 50:
            p.showPage()
            y = height - 50
        p.drawString(50, y, str(idx))
        p.drawString(80, y, row['person'].name)
        p.drawString(180, y, row['date'].strftime('%Y-%m-%d'))
        p.drawString(260, y, row['checkin'].strftime('%H:%M:%S') if row['checkin'] else '-')
        p.drawString(340, y, row['checkout'].strftime('%H:%M:%S') if row['checkout'] else '-')
        p.drawString(430, y, str(row['duration']) if row['duration'] else '-')
        y -= 18

    p.save()
    return response

# Export API (Attendance List)
@api_view(['GET'])
def attendance_list_api(request):
    records = Attendance.objects.select_related('person').order_by('-timestamp')
    serializer = AttendanceSerializer(records, many=True)
    return Response(serializer.data)


# Export API (Paired Attendance List)
@api_view(['GET'])
def paired_attendance_api(request):
    all_records = Attendance.objects.select_related('person').order_by('person_id', 'date', 'timestamp')

    grouped = defaultdict(list)
    for record in all_records:
        key = (record.person.id, record.date)
        grouped[key].append(record)

    results = []

    for (person_id, date), records in grouped.items():
        records = sorted(records, key=lambda r: r.timestamp)
        pending_checkins = deque()

        for rec in records:
            if rec.action == 'in':
                pending_checkins.append(rec.timestamp)
            elif rec.action == 'out':
                if pending_checkins:
                    checkin_time = pending_checkins.popleft()
                    duration = rec.timestamp - checkin_time
                    results.append({
                        "name": rec.person.name,
                        "date": date.strftime('%Y-%m-%d'),
                        "checkin": checkin_time.strftime('%H:%M:%S'),
                        "checkout": rec.timestamp.strftime('%H:%M:%S'),
                        "duration": str(duration)
                    })
                else:
                    results.append({
                        "name": rec.person.name,
                        "date": date.strftime('%Y-%m-%d'),
                        "checkin": None,
                        "checkout": rec.timestamp.strftime('%H:%M:%S'),
                        "duration": None
                    })

        while pending_checkins:
            checkin_time = pending_checkins.popleft()
            results.append({
                "name": rec.person.name,
                "date": date.strftime('%Y-%m-%d'),
                "checkin": checkin_time.strftime('%H:%M:%S'),
                "checkout": None,
                "duration": None
            })

    return Response(results)

# API (Person Profile List)
class PersonProfileListCreateAPI(generics.ListCreateAPIView):
    queryset = PersonProfile.objects.all()
    serializer_class = PersonProfileSerializer

# Save Data when mobile device(API) is scanned
@api_view(['POST'])
def scan_qr_api(request):
    qr_data = request.data.get('qr_data')

    if not qr_data:
        return Response({"message": "QR data is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        pk = int(qr_data.split(",")[0].strip())
        person = PersonProfile.objects.get(pk=pk)
    except (PersonProfile.DoesNotExist, ValueError, IndexError):
        return Response({"message": "Invalid QR code."}, status=status.HTTP_404_NOT_FOUND)

    today = timezone.now().date()
    latest_record = Attendance.objects.filter(
        person=person,
        date=today
    ).order_by('-timestamp').first()

    if not latest_record or latest_record.action == 'out':
        action = 'in'
    else:
        action = 'out'

    Attendance.objects.create(
        person=person,
        date=today,
        timestamp=timezone.now(),
        action=action
    )

    return Response({"message": f'AA-{person.reg_no}၊ {person.name} is {"Check In" if action == "in" else "Check Out"} Scanned Successfully! '})


# Monthly Person Report
def monthly_person_report(request):
     # Today
    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    # Get date filter params
    start_str = request.GET.get("start_date")
    end_str = request.GET.get("end_date")

    if start_str and end_str:
        try:
            start_date = parser.parse(start_str).date()
            end_date = parser.parse(end_str).date()
        except ValueError:
            start_date = first_day_of_month
            end_date = last_day_of_month
    else:
        start_date = first_day_of_month
        end_date = last_day_of_month

    # Query records
    records = Attendance.objects.filter(
        date__range=(start_date, end_date)
    ).select_related("person").order_by("person_id", "timestamp")

    grouped = defaultdict(list)
    for r in records:
        grouped[r.person].append(r)

    report = []

    for person, recs in grouped.items():
        recs = sorted(recs, key=lambda r: r.timestamp)
        pending_checkins = deque()
        total_duration = timedelta(0)
        count = 0

        for r in recs:
            if r.action == "in":
                pending_checkins.append(r.timestamp)
            elif r.action == "out":
                if pending_checkins:
                    checkin_time = pending_checkins.popleft()
                    duration = r.timestamp - checkin_time
                    total_duration += duration
                    count += 1

        report.append({
            "person": person,
            "checkin_count": count,
            "total_duration": format_duration(total_duration)
        })

    return render(request, "attendance/monthly_report.html", {
        "report": report,
        "start_date": start_date,
        "end_date": end_date
    })

# Export Monthly Report Excel
def export_monthly_report_excel(request):
    # Same Date Range Logic
    start_str = request.GET.get("start_date")
    end_str = request.GET.get("end_date")

    if start_str and end_str:
        start_date = parser.parse(start_str).date()
        end_date = parser.parse(end_str).date()
    else:
        start_date = timezone.now().replace(day=1).date()
        end_date = timezone.now().date()
    records = Attendance.objects.filter(
        date__range=(start_date, end_date)
    ).select_related("person").order_by("person_id", "timestamp")

    grouped = defaultdict(list)
    for r in records:
        grouped[r.person].append(r)

    report = []

    for person, recs in grouped.items():
        recs = sorted(recs, key=lambda r: r.timestamp)
        pending_checkins = deque()
        total_duration = timedelta(0)
        count = 0

        for r in recs:
            if r.action == "in":
                pending_checkins.append(r.timestamp)
            elif r.action == "out":
                if pending_checkins:
                    checkin_time = pending_checkins.popleft()
                    duration = r.timestamp - checkin_time
                    total_duration += duration
                    count += 1

        report.append({
            "person": person,
            "checkin_count": count,
            "total_duration": format_duration(total_duration)
        })

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    ws.append(["No", "Name", "Number of Check-in", "Total Duration"])

    for idx, row in enumerate(report, start=1):
        ws.append([
            idx,
            row["person"].name,
            row["checkin_count"],
            row["total_duration"]
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=monthly_report.xlsx"
    wb.save(response)
    return response

# Monthly Person Detailed Report 
def monthly_person_detailed_report(request):
    # Today
    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    # Get date filter params
    start_str = request.GET.get("start_date")
    end_str = request.GET.get("end_date")

    if start_str and end_str:
        try:
            start_date = parser.parse(start_str).date()
            end_date = parser.parse(end_str).date()
        except ValueError:
            start_date = first_day_of_month
            end_date = last_day_of_month
    else:
        start_date = first_day_of_month
        end_date = last_day_of_month

    # Query records
    records = Attendance.objects.filter(
        date__range=(start_date, end_date)
    ).select_related("person").order_by("person_id", "timestamp")

    grouped = defaultdict(list)
    for r in records:
        grouped[r.person].append(r)

    report = []

    for person, recs in grouped.items():
        recs = sorted(recs, key=lambda r: r.timestamp)
        pending_checkins = deque()
        detailed_rows = []
        total_duration = timedelta(0)

        for r in recs:
            if r.action == "in":
                pending_checkins.append(r.timestamp)
            elif r.action == "out":
                if pending_checkins:
                    checkin_time = pending_checkins.popleft()
                    duration = r.timestamp - checkin_time
                    total_duration += duration
                    detailed_rows.append({
                        "checkin": checkin_time,
                        "checkout": r.timestamp,
                        "duration": format_duration(duration)
                    })

        report.append({
            "person": person,
            "rows": detailed_rows,
            "total_duration": format_duration(total_duration),
            "count": len(detailed_rows)
        })

    return render(request, "attendance/monthly_detailed_report.html", {
        "report": report,
        "start_date": start_date,
        "end_date": end_date
    })


# Export Monthly Detailed Excel
def export_monthly_detailed_excel(request):
    # Date Range
    start_str = request.GET.get("start_date")
    end_str = request.GET.get("end_date")

    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    if start_str and end_str:
        start_date = parser.parse(start_str).date()
        end_date = parser.parse(end_str).date()
    else:
        start_date = first_day_of_month
        end_date = last_day_of_month

    records = Attendance.objects.filter(
        date__range=(start_date, end_date)
    ).select_related("person").order_by("person_id", "timestamp")

    grouped = defaultdict(list)
    for r in records:
        grouped[r.person].append(r)

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    row_num = 1
    ws.append(["Person", "Check-in", "Checkout", "Duration"])

    for person, recs in grouped.items():
        recs = sorted(recs, key=lambda r: r.timestamp)
        pending_checkins = deque()

        for r in recs:
            if r.action == "in":
                pending_checkins.append(r.timestamp)
            elif r.action == "out":
                if pending_checkins:
                    checkin_time = pending_checkins.popleft()
                    duration = r.timestamp - checkin_time
                    ws.append([
                        person.name,
                        checkin_time.strftime("%Y-%m-%d %H:%M:%S"),
                        r.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                        format_duration(duration)
                    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=monthly_detailed_report.xlsx"
    wb.save(response)
    return response


# Export Monthly Detailed PDF
def export_monthly_detailed_pdf(request):
    start_str = request.GET.get("start_date")
    end_str = request.GET.get("end_date")

    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    if start_str and end_str:
        start_date = parser.parse(start_str).date()
        end_date = parser.parse(end_str).date()
    else:
        start_date = first_day_of_month
        end_date = last_day_of_month

    records = Attendance.objects.filter(
        date__range=(start_date, end_date)
    ).select_related("person").order_by("person_id", "timestamp")

    grouped = defaultdict(list)
    for r in records:
        grouped[r.person].append(r)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=monthly_detailed_report.pdf"

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 50

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Monthly Detailed Attendance Report")
    y -= 30

    p.setFont("Helvetica", 10)

    for person, recs in grouped.items():
        p.drawString(50, y, f"Person: {person.name}")
        y -= 20
        p.drawString(70, y, "Check-in                      Checkout                      Duration")
        y -= 15

        recs = sorted(recs, key=lambda r: r.timestamp)
        pending_checkins = deque()

        for r in recs:
            if r.action == "in":
                pending_checkins.append(r.timestamp)
            elif r.action == "out":
                if pending_checkins:
                    checkin_time = pending_checkins.popleft()
                    duration = r.timestamp - checkin_time
                    p.drawString(70, y, f"{checkin_time.strftime('%Y-%m-%d %H:%M:%S')}   {r.timestamp.strftime('%Y-%m-%d %H:%M:%S')}   {format_duration(duration)}")
                    y -= 15
                    if y < 50:
                        p.showPage()
                        y = height - 50

        y -= 20

    p.showPage()
    p.save()
    return response
